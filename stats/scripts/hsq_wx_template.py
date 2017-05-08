import os
import sys

import arrow
from openpyxl import load_workbook, Workbook

READ_PATH = '{root_dir}/{file_dir}/{file_name}'.format(
                root_dir=os.path.abspath(os.path.dirname(__file__)),
                file_dir='files',
                file_name='wxtemplate.xlsx'
            )
WRITE_PATH = '{root_dir}/{file_dir}/{file_name}'.format(
                root_dir=os.path.abspath(os.path.dirname(__file__)),
                file_dir='files',
                file_name='wxtemplate_result.xlsx'
            )
EXCLUDE = ('brief', 'template')
WS_KEYS = ('类型', '开始时间', '结束时间', '说明', 'sql', '条件')
TYPES = ('拼团', '抽奖', '领券', '其他')


def get_sheets(wb):
    wb = load_workbook(READ_PATH)
    return wb.get_sheet_names()


def get_info(ws, cnx):
    query_data = {}
    for i in range(len(WS_KEYS)):
        if WS_KEYS[i] != ws['A{}'.format(i+1)].value:
            print(ws['A{}'.format(i+1)].value)
            raise RuntimeError(
                'format error! sth wrong with "{}"'.format(WS_KEYS[i]))
        else:
            if i < 5:
                query_data[WS_KEYS[i]] = ws['B{}'.format(i+1)].value
            else:
                query_data['条件'] = get_conditions(ws, query_data['类型'], cnx)

    if query_data['类型'] not in TYPES:
        raise RuntimeError(
            'format error! not valid type: {}'.format(query_data['类型']))

    try:
        if query_data['开始时间']:
            query_data['开始时间'] = arrow.get(query_data['开始时间'])
        if query_data['结束时间']:
            query_data['结束时间'] = arrow.get(query_data['结束时间'])
    except:
        raise RuntimeError('format error! please check start time/end time')

    # if query_data['sql']\
    #    and (not str(query_data['sql']).startswith('select') or
    #         not str(query_data['sql']).startswith('SELECT')):
    #     raise RuntimeError(
    #         "format error! invalid sql with {}".format(query_data['sql']))

    return query_data


def get_conditions(ws, qtype, cnx):
    vals = []
    if qtype in ('拼团', '抽奖', '领券'):
        for i in range(5, ws.max_row):
            if ws.cell(row=i+1, column=2).value:
                try:
                    vals.append(str(int(ws.cell(row=i+1, column=2).value)))
                except:
                    raise RuntimeError(
                        'format error! not valid condition found')

    if qtype == '领券':
        coupon_ids = []
        sql = '''
                select distinct(coupon_id)
                from reward_rules
                where reward_id in ({aids})
        '''.format(aids=','.join(vals))

        cursor = cnx.cursor()
        cursor.execute(sql)
        coupons = cursor.fetchall()
        cursor.close()

        for (coupon_id,) in coupons:
            coupon_ids.append(str(coupon_id))

        vals = coupon_ids

    return vals


def generate_sql(qd, exclude_users):
    sql = ''

    if qd['类型'] == '拼团':
        sql = '''
                select distinct(o.user_id), u.wechat_mp_open_id,
                       o.pin_event_status status
                from pin_activities_order o
                inner join user_login_info u on u.user_id=o.user_id
                where pin_activities_id in ({pids})
                and (u.wechat_mp_open_id is not null
                     and u.wechat_mp_open_id != '')
        '''.format(pids=','.join(qd['条件']))
    elif qd['类型'] == '抽奖':
        sql = '''
                select distinct(o.user_id), u.wechat_mp_open_id,
                       if(o.pin_event_status=3,-1,o.is_winner) status
                from pin_activities_order o
                inner join user_login_info u on u.user_id=o.user_id
                where pin_activities_id in ({pids})
                and o.user_id not in (
                select user_id from (
                    select user_id, count(0) cnt, sum(is_winner) win
                    from pin_activities_order o
                    inner join pin_activities a on o.pin_activities_id=a.id
                    where a.type=2
                    group by user_id)t where cnt>9 and win=0
                )
                and (u.wechat_mp_open_id is not null
                     and u.wechat_mp_open_id != '')
        '''.format(pids=','.join(qd['条件']))
    elif qd['类型'] == '领券':
        sql = '''
                select distinct(o.user_id), ul.wechat_mp_open_id
                from trade_order_coupon o
                inner join `user_login_info` ul on ul.user_id = o.user_id
                where o.coupon_id in ({cids})
                and (ul.wechat_mp_open_id is not null
                     and ul.wechat_mp_open_id != '')
        '''.format(cids=','.join(qd['条件']))
    elif qd['类型'] == '其他':
        sql = qd['sql']

    if exclude_users:
        sql += " and o.user_id not in ({exist_users})".format(
            exist_users=','.join(exclude_users))
    if qd['类型'] != '其他' and qd['开始时间']:
        sql += " and created_at>unix_timestamp('{}')".format(
            qd['开始时间'].format('YYYY-MM-DD'))
    if qd['类型'] != '其他' and qd['结束时间']:
        sql += " and created_at<unix_timestamp('{} 23:59:59')".format(
            qd['结束时间'].format('YYYY-MM-DD'))

    return sql


def get_users(cnx, sqlstr):
    cursor = cnx.cursor()
    cursor.execute(sqlstr)
    users = cursor.fetchall()
    cursor.close()
    return users


def write_results(worksheet, qd_type, users, selected_users):
    if qd_type in ('领券', '其他'):
        worksheet['A1'] = 'User_ID'
        worksheet['B1'] = 'Open_ID'
        for i, user in enumerate(users, 2):
            selected_users.append(str(user[0]))
            worksheet['A{}'.format(i)] = user[0]
            worksheet['B{}'.format(i)] = user[1]
    if qd_type == '拼团':
        worksheet['A1'] = 'User_ID [拼团失败]'
        worksheet['B1'] = 'Open_ID [拼团失败]'
        worksheet['D1'] = 'User_ID [拼团成功]'
        worksheet['E1'] = 'Open_ID [拼团成功]'
        row_index = [2, 2]
        for user in users:
            selected_users.append(str(user[0]))
            if user[-1] == 3:
                worksheet['A{}'.format(row_index[0])] = user[0]
                worksheet['B{}'.format(row_index[0])] = user[1]
                row_index[0] += 1
            if user[-1] == 2:
                worksheet['D{}'.format(row_index[1])] = user[0]
                worksheet['E{}'.format(row_index[1])] = user[1]
                row_index[1] += 1
    if qd_type == '抽奖':
        worksheet['A1'] = 'User_ID [拼团失败]'
        worksheet['B1'] = 'Open_ID [拼团失败]'
        worksheet['D1'] = 'User_ID [未中奖]'
        worksheet['E1'] = 'Open_ID [未中奖]'
        worksheet['G1'] = 'User_ID [已中奖]'
        worksheet['H1'] = 'Open_ID [已中奖]'
        row_index = [2, 2, 2]
        for user in users:
            selected_users.append(str(user[0]))
            if user[-1] == -1:
                worksheet['A{}'.format(row_index[0])] = user[0]
                worksheet['B{}'.format(row_index[0])] = user[1]
                row_index[0] += 1
            if user[-1] == 0:
                worksheet['D{}'.format(row_index[1])] = user[0]
                worksheet['E{}'.format(row_index[1])] = user[1]
                row_index[1] += 1
            if user[-1] == 1:
                worksheet['G{}'.format(row_index[2])] = user[0]
                worksheet['H{}'.format(row_index[2])] = user[1]
                row_index[2] += 1


if __name__ == '__main__':
    import_path = '{}/../../'.format(
        os.path.abspath(os.path.dirname(__file__)))
    sys.path.append(import_path)

    from stats.models.mysql import init_mysql

    try:
        cnx = init_mysql('hsq_ro')
        result_wb = Workbook()
        selected_users = []
        wb = load_workbook(READ_PATH)
        ws_list = get_sheets(wb)
        for ws_name in ws_list:
            if ws_name not in EXCLUDE:
                print('working on worksheet {} ...'.format(ws_name))
                ws = wb.get_sheet_by_name(ws_name)
                print('    checking format ...')
                qd = get_info(ws, cnx)
                qd['sql'] = generate_sql(qd, selected_users)

                print('    getting user info ...')
                users = get_users(cnx, qd['sql'])

                print('    writing to file ...')
                result_ws = result_wb.create_sheet(ws_name)
                write_results(result_ws, qd['类型'], users, selected_users)

        result_wb.save(WRITE_PATH)
        print('select {} users totally'.format(len(selected_users)))
        print('done!')
    except Exception as e:
        print(e)
    finally:
        cnx.close()
